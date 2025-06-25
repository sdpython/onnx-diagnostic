import datetime
import enum
import glob
import io
import os
import pprint
import re
import warnings
import zipfile
from typing import Any, Callable, Dict, Iterator, List, Optional, Sequence, Tuple, Union
import numpy as np
import pandas
from pandas.api.types import is_numeric_dtype, is_datetime64_any_dtype
from .helper import string_sig

BUCKET_SCALES_VALUES = np.array(
    [-np.inf, -20, -10, -5, -2, 0, 2, 5, 10, 20, 100, 200, 300, 400, np.inf], dtype=float
)


BUCKET_SCALES = BUCKET_SCALES_VALUES / 100 + 1


def filter_data(
    df: pandas.DataFrame,
    filter_in: Optional[str] = None,
    filter_out: Optional[str] = None,
    verbose: int = 0,
) -> pandas.DataFrame:
    """
    Argument `filter` follows the syntax
    ``<column1>:<fmt1>//<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    if not filter_in and not filter_out:
        return df

    def _f(fmt):
        cond = {}
        if isinstance(fmt, str):
            cols = fmt.split("//")
            for c in cols:
                assert ":" in c, f"Unexpected value {c!r} in fmt={fmt!r}"
                spl = c.split(":")
                assert len(spl) == 2, f"Unexpected value {c!r} in fmt={fmt!r}"
                name, fil = spl
                cond[name] = set(fil.split(";"))
        return cond

    if filter_in:
        cond = _f(filter_in)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_in!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            if verbose:
                print(
                    f"[_filter_data] filter in column {k!r}, "
                    f"values {v!r} among {set(df[k].astype(str))}"
                )
            df = df[df[k].astype(str).isin(v)]

    if filter_out:
        cond = _f(filter_out)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_out!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            if verbose:
                print(
                    f"[_filter_data] filter out column {k!r}, "
                    f"values {v!r} among {set(df[k].astype(str))}"
                )
            df = df[~df[k].astype(str).isin(v)]
    return df


def enumerate_csv_files(
    data: Union[
        pandas.DataFrame, List[Union[str, Tuple[str, str]]], str, Tuple[str, str, str, str]
    ],
    verbose: int = 0,
    filtering: Optional[Callable[[str], bool]] = None,
) -> Iterator[Union[pandas.DataFrame, str, Tuple[str, str, str, str]]]:
    """
    Enumerates files considered for the aggregation.
    Only csv files are considered.
    If a zip file is given, the function digs into the zip files and
    loops over csv candidates.

    :param data: dataframe with the raw data or a file or list of files
    :param vrbose: verbosity
    :param filtering: function to filter in or out files in zip files,
        must return true to keep the file, false to skip it.
    :return: a generator yielding tuples with the filename, date, full path and zip file

    data can contains:
    * a dataframe
    * a string for a filename, zip or csv
    * a list of string
    * a tuple
    """
    if not isinstance(data, list):
        data = [data]
    for itn, filename in enumerate(data):
        if isinstance(filename, pandas.DataFrame):
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}] is a dataframe")
            yield filename
            continue

        if isinstance(filename, tuple):
            # A file in a zipfile
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}] is {filename!r}")
            yield filename
            continue

        if os.path.exists(filename):
            ext = os.path.splitext(filename)[-1]
            if ext == ".csv":
                # We check the first line is ok.
                if verbose:
                    print(f"[enumerate_csv_files] data[{itn}] is a csv file: {filename!r}]")
                dt = datetime.datetime.fromtimestamp(os.stat(filename).st_mtime)
                du = dt.strftime("%Y-%m-%d %H:%M:%S")
                yield (os.path.split(filename)[-1], du, filename, "")
                continue

            if ext == ".zip":
                if verbose:
                    print(f"[enumerate_csv_files] data[{itn}] is a zip file: {filename!r}]")
                zf = zipfile.ZipFile(filename, "r")
                for ii, info in enumerate(zf.infolist()):
                    name = info.filename
                    if filtering is None:
                        ext = os.path.splitext(name)[-1]
                        if ext != ".csv":
                            continue
                    elif not filtering(name):
                        continue
                    if verbose:
                        print(
                            f"[enumerate_csv_files] data[{itn}][{ii}] is a csv file: {name!r}]"
                        )
                    with zf.open(name) as zzf:
                        first_line = zzf.readline()
                    if b"," not in first_line:
                        continue
                    yield (
                        os.path.split(name)[-1],
                        "%04d-%02d-%02d %02d:%02d:%02d" % info.date_time,
                        name,
                        filename,
                    )
                zf.close()
                continue

            raise AssertionError(f"Unexpected format {filename!r}, cannot read it.")

        # filename is a pattern.
        found = glob.glob(filename)
        if verbose and not found:
            print(f"[enumerate_csv_files] unable to find file in {filename!r}")
        for ii, f in enumerate(found):
            if verbose:
                print(f"[enumerate_csv_files] data[{itn}][{ii}] {f!r} from {filename!r}")
            yield from enumerate_csv_files(f, verbose=verbose, filtering=filtering)


def open_dataframe(
    data: Union[str, Tuple[str, str, str, str], pandas.DataFrame],
) -> pandas.DataFrame:
    """
    Opens a filename defined by function
    :func:`onnx_diagnostic.helpers.log_helper.enumerate_csv_files`.

    :param data: a dataframe, a filename, a tuple indicating the file is coming
        from a zip file
    :return: a dataframe
    """
    if isinstance(data, pandas.DataFrame):
        return data
    if isinstance(data, str):
        df = pandas.read_csv(data)
        df["RAWFILENAME"] = data
        return df
    if isinstance(data, tuple):
        if not data[-1]:
            df = pandas.read_csv(data[2])
            df["RAWFILENAME"] = data[2]
            return df
        zf = zipfile.ZipFile(data[-1])
        with zf.open(data[2]) as f:
            df = pandas.read_csv(f)
            df["RAWFILENAME"] = f"{data[-1]}/{data[2]}"
        zf.close()
        return df

    raise ValueError(f"Unexpected value for data: {data!r}")


class CubeViewDef:
    """
    Defines how to compute a view.

    :param key_index: keys to put in the row index
    :param values: values to show
    :param ignore_unique: ignore keys with a unique value
    :param order: to reorder key in columns index
    :param key_agg: aggregate according to these columns before
        creating the view
    :param agg_args: see :meth:`pandas.core.groupby.DataFrameGroupBy.agg`,
        it can be also a callable to return a different aggregation
        method depending on the column name
    :param agg_kwargs: see :meth:`pandas.core.groupby.DataFrameGroupBy.agg`
    :param agg_multi: aggregation over multiple columns
    :param ignore_columns: ignore the following columns if known to overload the view
    :param keep_columns_in_index: keeps the columns even if there is only one unique value
    :param dropna: drops rows with nan if not relevant
    :param transpose: transpose
    :param f_highlight: to highlights some values
    :param name: name of the view, used mostly to debug
    :param plots: adds plot to the Excel sheet
    :param no_index: remove the index (but keeps the columns)
    """

    class HighLightKind(enum.IntEnum):
        NONE = 0
        RED = 1
        GREEN = 2

    def __init__(
        self,
        key_index: Sequence[str],
        values: Sequence[str],
        ignore_unique: bool = True,
        order: Optional[Sequence[str]] = None,
        key_agg: Optional[Sequence[str]] = None,
        agg_args: Union[Sequence[Any], Callable[[str], Any]] = ("sum",),
        agg_kwargs: Optional[Dict[str, Any]] = None,
        agg_multi: Optional[
            Dict[str, Callable[[pandas.core.groupby.DataFrameGroupBy], pandas.Series]]
        ] = None,
        ignore_columns: Optional[Sequence[str]] = None,
        keep_columns_in_index: Optional[Sequence[str]] = None,
        dropna: bool = True,
        transpose: bool = False,
        f_highlight: Optional[Callable[[Any], "CubeViewDef.HighLightKind"]] = None,
        name: Optional[str] = None,
        no_index: bool = False,
        plots: bool = False,
    ):
        self.key_index = key_index
        self.values = values
        self.ignore_unique = ignore_unique
        self.order = order
        self.key_agg = key_agg
        self.agg_args = agg_args
        self.agg_kwargs = agg_kwargs
        self.agg_multi = agg_multi
        self.dropna = dropna
        self.ignore_columns = ignore_columns
        self.keep_columns_in_index = keep_columns_in_index
        self.f_highlight = f_highlight
        self.transpose = transpose
        self.name = name
        self.no_index = no_index
        self.plots = plots

    def __repr__(self) -> str:
        "usual"
        return string_sig(self)  # type: ignore[arg-type]


def apply_excel_style(
    filename_or_writer: Any,
    f_highlights: Optional[Dict[str, Callable[[Any], CubeViewDef.HighLightKind]]] = None,
):
    """
    Applies styles on all sheets in a file unless the sheet is too big.

    :param filename_or_writer: filename, modified inplace
    :param f_highlight: color function to apply, one per sheet
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font  # , PatternFill, numbers

    if isinstance(filename_or_writer, str):
        workbook = load_workbook(filename_or_writer)
        save = True
    else:
        workbook = filename_or_writer.book
        save = False

    left = Alignment(horizontal="left")
    left_shrink = Alignment(horizontal="left", shrink_to_fit=True)
    right = Alignment(horizontal="right")
    font_colors = {
        CubeViewDef.HighLightKind.GREEN: Font(color="00AA00"),
        CubeViewDef.HighLightKind.RED: Font(color="FF0000"),
    }

    for name in workbook.sheetnames:
        f_highlight = f_highlights.get(name, None) if f_highlights else None
        sheet = workbook[name]
        n_rows = sheet.max_row
        n_cols = sheet.max_column
        if n_rows * n_cols > 2**18:
            # Too big.
            continue
        co: Dict[int, int] = {}
        sizes: Dict[int, int] = {}
        cols = set()
        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                cols.add(cell.column)
                if isinstance(cell.value, float):
                    co[j] = co.get(j, 0) + 1
                elif isinstance(cell.value, str):
                    sizes[cell.column] = max(sizes.get(cell.column, 0), len(cell.value))

        for k, v in sizes.items():
            c = get_column_letter(k)
            sheet.column_dimensions[c].width = min(max(8, v), 30)
        for k in cols:
            if k not in sizes:
                c = get_column_letter(k)
                sheet.column_dimensions[c].width = 15

        for i in range(1, n_rows + 1):
            for j, cell in enumerate(sheet[i]):
                if j > n_cols:
                    break
                if isinstance(cell.value, pandas.Timestamp):
                    cell.alignment = right
                    dt = cell.value.to_pydatetime()
                    cell.value = dt
                    cell.number_format = (
                        "YYYY-MM-DD"
                        if (
                            dt.hour == 0
                            and dt.minute == 0
                            and dt.second == 0
                            and dt.microsecond == 0
                        )
                        else "YYYY-MM-DD 00:00:00"
                    )
                elif isinstance(cell.value, (float, int)):
                    cell.alignment = right
                    x = abs(cell.value)
                    if int(x) == x:
                        cell.number_format = "0"
                    elif x > 5000:
                        cell.number_format = "# ##0"
                    elif x >= 500:
                        cell.number_format = "0.0"
                    elif x >= 50:
                        cell.number_format = "0.00"
                    elif x >= 5:
                        cell.number_format = "0.000"
                    elif x > 0.5:
                        cell.number_format = "0.0000"
                    elif x > 0.005:
                        cell.number_format = "0.00000"
                    else:
                        cell.number_format = "0.000E+00"
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]
                elif isinstance(cell.value, str) and len(cell.value) > 70:
                    cell.alignment = left_shrink
                else:
                    cell.alignment = left
                    if f_highlight:
                        h = f_highlight(cell.value)
                        if h in font_colors:
                            cell.font = font_colors[h]
    if save:
        workbook.save(filename_or_writer)


class CubePlot:
    """
    Creates a plot.

    :param df: dataframe
    :param kind: kind of graph to plot, bar, barh, line
    :param split: draw a graph per line in the dataframe
    :param timeseries: this assumes the time is one level of the columns,
        this argument indices the level name
    """

    KINDS = {"bar", "barh", "line"}

    @classmethod
    def group_columns(
        cls, columns: List[str], sep: str = "/", depth: int = 2
    ) -> List[List[str]]:
        """Groups columns to have nice display."""
        res: Dict[str, List[str]] = {}
        for c in columns:
            p = c.split("/")
            k = "/".join(p[:depth])
            if k not in res:
                res[k] = []
            res[k].append(c)
        new_res: Dict[str, List[str]] = {}
        for k, v in res.items():
            if len(v) >= 3:
                new_res[k] = v
            else:
                if "0" not in new_res:
                    new_res["0"] = []
                new_res["0"].extend(v)
        groups: List[List[str]] = [sorted(v) for k, v in sorted(new_res.items())]
        if depth <= 1:
            return groups
        new_groups: List[List[str]] = []
        for v in groups:
            if len(v) >= 6:
                new_groups.extend(cls.group_columns(v, depth=1, sep=sep))
            else:
                new_groups.append(v)
        return new_groups

    def __init__(
        self,
        df: pandas.DataFrame,
        kind: str = "bar",
        orientation="col",
        split: bool = True,
        timeseries: Optional[str] = None,
    ):
        assert (
            not timeseries or timeseries in df.columns.names
        ), f"Level {timeseries!r} is not part of the columns levels {df.columns.names}"
        assert (
            kind in self.__class__.KINDS
        ), f"Unexpected kind={kind!r} not in {self.__class__.KINDS}"
        assert split, f"split={split} not implemented"
        assert (
            not timeseries or orientation == "row"
        ), f"orientation={orientation!r} must be 'row' for timeseries"
        self.df = df.copy()
        self.kind = kind
        self.orientation = orientation
        self.split = split
        self.timeseries = timeseries

        if timeseries:
            if isinstance(self.df.columns, pandas.MultiIndex):
                index_time = list(self.df.columns.names).index(self.timeseries)

                def _drop(t, i=index_time):
                    return (*t[:i], *t[i + 1 :])

                self.df.columns = pandas.MultiIndex.from_tuples(
                    [("/".join(map(str, _drop(i))), i[index_time]) for i in self.df.columns],
                    names=["metric", timeseries],
                )
        else:
            if isinstance(self.df.columns, pandas.MultiIndex):
                self.df.columns = ["/".join(map(str, i)) for i in self.df.columns]
        if isinstance(self.df.index, pandas.MultiIndex):
            self.df.index = ["/".join(map(str, i)) for i in self.df.index]

    def __repr__(self) -> str:
        "usual"
        return string_sig(self)  # type: ignore[arg-type]

    def to_images(
        self, verbose: int = 0, merge: bool = True, title_suffix: Optional[str] = None
    ) -> List[bytes]:
        """
        Converts data into plots and images.

        :param verbose: verbosity
        :param merge: returns all graphs in a single image (True)
            or an image for every graph (False)
        :param title_suffix: prefix for the title of every graph
        :return: list of binary images (format PNG)
        """
        if self.kind in ("barh", "bar"):
            return self._to_images_bar(verbose=verbose, merge=merge, title_suffix=title_suffix)
        if self.kind == "line":
            return self._to_images_line(
                verbose=verbose, merge=merge, title_suffix=title_suffix
            )
        raise AssertionError(f"self.kind={self.kind!r} not implemented")

    @classmethod
    def _make_loop(cls, ensemble, verbose):
        if verbose:
            from tqdm import tqdm

            loop = tqdm(ensemble)
        else:
            loop = ensemble
        return loop

    def _to_images_bar(
        self, verbose: int = 0, merge: bool = True, title_suffix: Optional[str] = None
    ) -> List[bytes]:
        assert merge, f"merge={merge} not implemented yet"
        import matplotlib.pyplot as plt

        df = self.df.T if self.orientation == "row" else self.df
        title_suffix = f"\n{title_suffix}" if title_suffix else ""

        n_cols = 3
        nn = df.shape[1] // n_cols
        nn += int(df.shape[1] % n_cols != 0)
        fig, axs = plt.subplots(nn, n_cols, figsize=(6 * n_cols, nn * df.shape[0] / 5))
        pos = 0
        imgs = []
        for c in self._make_loop(df.columns, verbose):
            ax = axs[pos // n_cols, pos % n_cols]
            (
                df[c].plot.barh(title=f"{c}{title_suffix}", ax=ax)
                if self.kind == "barh"
                else df[c].plot.bar(title=f"{c}{title_suffix}", ax=ax)
            )
            ax.tick_params(axis="both", which="major", labelsize=8)
            ax.grid(True)
            pos += 1  # noqa: SIM113
        fig.tight_layout()
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format="png")
        imgs.append(imgdata.getvalue())
        plt.close()
        return imgs

    def _to_images_line(
        self, verbose: int = 0, merge: bool = True, title_suffix: Optional[str] = None
    ) -> List[bytes]:
        assert merge, f"merge={merge} not implemented yet"
        assert (
            self.orientation == "row"
        ), f"self.orientation={self.orientation!r} not implemented for this kind of graph."

        def rotate_align(ax, angle=15, align="right"):
            for label in ax.get_xticklabels():
                label.set_rotation(angle)
                label.set_horizontalalignment(align)
            ax.tick_params(axis="both", which="major", labelsize=8)
            ax.grid(True)
            ax.legend()
            ax.tick_params(labelleft=True)
            return ax

        import matplotlib.pyplot as plt

        df = self.df.T

        confs = list(df.unstack(self.timeseries).index)
        groups = self.group_columns(confs)
        n_cols = len(groups)

        title_suffix = f"\n{title_suffix}" if title_suffix else ""
        fig, axs = plt.subplots(
            df.shape[1],
            n_cols,
            figsize=(5 * n_cols, max(len(g) for g in groups) * df.shape[1] / 2),
            sharex=True,
            sharey="row" if n_cols > 1 else False,
        )
        imgs = []
        row = 0
        for c in self._make_loop(df.columns, verbose):
            dfc = df[[c]]
            dfc = dfc.unstack(self.timeseries).T.droplevel(0)
            if n_cols == 1:
                dfc.plot(title=f"{c}{title_suffix}", ax=axs[row], linewidth=3)
                axs[row].grid(True)
                rotate_align(axs[row])
            else:
                x = list(range(dfc.shape[0]))
                ticks = list(dfc.index)
                for ii, group in enumerate(groups):
                    ddd = dfc.loc[:, group].copy()
                    axs[row, ii].set_xticks(x)
                    axs[row, ii].set_xticklabels(ticks)
                    # This is very slow
                    # ddd.plot(ax=axs[row, ii],linewidth=3)
                    for jj in range(ddd.shape[1]):
                        axs[row, ii].plot(x, ddd.iloc[:, jj], lw=3, label=ddd.columns[jj])
                    axs[row, ii].set_title(f"{c}{title_suffix}")
                    rotate_align(axs[row, ii])
            row += 1  # noqa: SIM113
        fig.tight_layout()
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format="png")
        imgs.append(imgdata.getvalue())
        plt.close()
        return imgs


class CubeLogs:
    """
    Processes logs coming from experiments.
    """

    def __init__(
        self,
        data: Any,
        time: str = "date",
        keys: Sequence[str] = ("version_.*", "model_.*"),
        values: Sequence[str] = ("time_.*", "disc_.*"),
        ignored: Sequence[str] = (),
        recent: bool = False,
        formulas: Optional[
            Union[
                Sequence[str],
                Dict[str, Union[str, Callable[[pandas.DataFrame], pandas.Series]]],
            ]
        ] = None,
        fill_missing: Optional[Sequence[Tuple[str, Any]]] = None,
        keep_last_date: bool = False,
    ):
        self._data = data
        self._time = time
        self._keys = keys
        self._values = values
        self._ignored = ignored
        self.recent = recent
        self._formulas = formulas
        self.fill_missing = fill_missing
        self.keep_last_date = keep_last_date

    def post_load_process_piece(
        self, df: pandas.DataFrame, unique: bool = False
    ) -> pandas.DataFrame:
        """
        Postprocesses a piece when a cube is made of multiple pieces
        before it gets merged.
        """
        if not self.fill_missing:
            return df
        missing = dict(self.fill_missing)
        for k, v in missing.items():
            if k not in df.columns:
                df[k] = v
        return df

    def load(self, verbose: int = 0):
        """Loads and preprocesses the data. Returns self."""
        if isinstance(self._data, pandas.DataFrame):
            if verbose:
                print(f"[CubeLogs.load] load from dataframe, shape={self._data.shape}")
            self.data = self.post_load_process_piece(self._data, unique=True)
            if verbose:
                print(f"[CubeLogs.load] after postprocessing shape={self.data.shape}")
        elif isinstance(self._data, list) and all(isinstance(r, dict) for r in self._data):
            if verbose:
                print(f"[CubeLogs.load] load from list of dicts, n={len(self._data)}")
            self.data = pandas.DataFrame(self.post_load_process_piece(self._data, unique=True))
            if verbose:
                print(f"[CubeLogs.load] after postprocessing shape={self.data.shape}")
        elif isinstance(self._data, list) and all(
            isinstance(r, pandas.DataFrame) for r in self._data
        ):
            if verbose:
                print(f"[CubeLogs.load] load from list of DataFrame, n={len(self._data)}")
            self.data = pandas.concat(
                [self.post_load_process_piece(c) for c in self._data], axis=0
            )
            if verbose:
                print(f"[CubeLogs.load] after postprocessing shape={self.data.shape}")
        elif isinstance(self._data, list):
            if verbose:
                print("[CubeLogs.load] load from list of Cubes")
            cubes = []
            for item in enumerate_csv_files(self._data, verbose=verbose):
                df = open_dataframe(item)
                cube = CubeLogs(
                    df,
                    time=self._time,
                    keys=self._keys,
                    values=self._values,
                    ignored=self._ignored,
                    recent=self.recent,
                )
                cube.load()
                cubes.append(self.post_load_process_piece(cube.data))
            self.data = pandas.concat(cubes, axis=0)
            if verbose:
                print(f"[CubeLogs.load] after postprocessing shape={self.data.shape}")
        else:
            raise NotImplementedError(
                f"Not implemented with the provided data (type={type(self._data)})"
            )

        assert all(isinstance(c, str) for c in self.data.columns), (
            f"The class only supports string as column names "
            f"but found {[c for c in self.data.columns if not isinstance(c, str)]}"
        )
        if verbose:
            print(f"[CubeLogs.load] loaded with shape={self.data.shape}")

        self._initialize_columns()
        if verbose:
            print(f"[CubeLogs.load] time={self.time}")
            print(f"[CubeLogs.load] keys={self.keys_no_time}")
            print(f"[CubeLogs.load] values={self.values}")
            print(f"[CubeLogs.load] ignored={self.ignored}")
            print(f"[CubeLogs.load] ignored_values={self.ignored_values}")
            print(f"[CubeLogs.load] ignored_keys={self.ignored_keys}")
        assert self.keys_no_time, f"No keys found with {self._keys} from {self.data.columns}"
        assert self.values, f"No values found with {self._values} from {self.data.columns}"
        assert not (
            set(self.keys_no_time) & set(self.values)
        ), f"Columns {set(self.keys_no_time) & set(self.values)} cannot be keys and values"
        assert not (
            set(self.keys_no_time) & set(self.ignored)
        ), f"Columns {set(self.keys_no_time) & set(self.ignored)} cannot be keys and ignored"
        assert not (
            set(self.values) & set(self.ignored)
        ), f"Columns {set(self.keys_no_time) & set(self.ignored)} cannot be values and ignored"
        assert (
            self.time not in self.keys_no_time
            and self.time not in self.values
            and self.time not in self.ignored
        ), (
            f"Column {self.time!r} is also a key, a value or ignored, "
            f"keys={sorted(self.keys_no_time)}, values={sorted(self.values)}, "
            f"ignored={sorted(self.ignored)}"
        )
        self._columns = [self.time, *self.keys_no_time, *self.values, *self.ignored]
        self.dropped = [c for c in self.data.columns if c not in set(self.columns)]
        self.data = self.data[self.columns]
        if verbose:
            print(f"[CubeLogs.load] dropped={self.dropped}")
            print(f"[CubeLogs.load] data.shape={self.data.shape}")

        shape = self.data.shape
        if verbose:
            print(f"[CubeLogs.load] removed columns, shape={self.data.shape}")
        self._preprocess()
        if verbose:
            print(f"[CubeLogs.load] preprocess, shape={self.data.shape}")
        assert (
            self.data.shape[0] > 0
        ), f"The preprocessing reduced shape {shape} to {self.data.shape}."
        if self.recent and verbose:
            print(f"[CubeLogs.load] keep most recent data.shape={self.data.shape}")

        # Let's apply the formulas
        if self._formulas:
            forms = (
                {k: k for k in self._formulas}
                if not isinstance(self._formulas, dict)
                else self._formulas
            )
            cols = set(self.values)
            for k, ff in forms.items():
                f = self._process_formula(ff)
                if k in cols or f is None:
                    if verbose:
                        print(f"[CubeLogs.load] skip formula {k!r}")
                else:
                    if verbose:
                        print(f"[CubeLogs.load] apply formula {k!r}")
                    self.data[k] = f(self.data)
                    self.values.append(k)
                    cols.add(k)
        self.values_for_key = {k: set(self.data[k].dropna()) for k in self.keys_time}
        for k in self.keys_no_time:
            if self.data[k].isna().max():
                self.values_for_key[k].add(np.nan)
        self.keys_with_nans = [
            c for c in self.keys_time if self.data[c].isna().astype(int).sum() > 0
        ]
        if verbose:
            print(f"[CubeLogs.load] convert column {self.time!r} into date")
            if self.keys_with_nans:
                print(f"[CubeLogs.load] keys_with_nans={self.keys_with_nans}")
        self.data[self.time] = pandas.to_datetime(self.data[self.time])

        if self.keep_last_date:
            times = self.data[self.time].dropna()
            mi, mx = times.min(), times.max()
            if mi != mx:
                print(f"[CubeLogs.load] setting all dates in column {self.time} to {mx!r}")
                self.data.loc[~self.data[self.time].isna(), self.time] = mx
                self.values_for_key[self.time] = {mx}
                if self.data[self.time].isna().max():
                    self.values_for_key[self.time].add(np.nan)
        if verbose:
            print(f"[CubeLogs.load] done, shape={self.shape}")
        return self

    def _process_formula(
        self, formula: Union[str, Callable[[pandas.DataFrame], pandas.Series]]
    ) -> Callable[[pandas.DataFrame], pandas.Series]:
        assert callable(formula), f"formula={formula!r} is not supported."
        return formula

    @property
    def shape(self) -> Tuple[int, int]:
        "Returns the shape."
        assert hasattr(self, "data"), "Method load was not called"
        return self.data.shape

    @property
    def columns(self) -> Sequence[str]:
        "Returns the columns."
        assert hasattr(self, "data"), "Method load was not called"
        return self.data.columns

    def _preprocess(self):
        last = self.values[0]
        gr = self.data[[*self.keys_time, last]].groupby(self.keys_time, dropna=False).count()
        gr = gr[gr[last] > 1]
        if self.recent:
            cp = self.data.copy()
            assert (
                "__index__" not in cp.columns
            ), f"'__index__' should not be a column in {cp.columns}"
            cp["__index__"] = np.arange(cp.shape[0])
            gr = (
                cp[[*self.keys_time, "__index__"]]
                .groupby(self.keys_no_time, as_index=False, dropna=False)
                .max()
            )
            assert gr.shape[0] > 0, (
                f"Something went wrong after the groupby.\n"
                f"{cp[[*self.keys, self.time, '__index__']].head().T}"
            )
            filtered = pandas.merge(cp, gr, on=["__index__", *self.keys_time])
            assert filtered.shape[0] <= self.data.shape[0], (
                f"Keeping the latest row brings more row {filtered.shape} "
                f"(initial is {self.data.shape})."
            )
            self.data = filtered.drop("__index__", axis=1)
        else:
            assert gr.shape[0] == 0, f"There are duplicated rows:\n{gr}"

    @classmethod
    def _filter_column(cls, filters, columns, can_be_empty=False):
        assert list(columns), "columns is empty"
        set_cols = set()
        for f in filters:
            if set(f) & {'"', "^", ".", "*", "+", "{", "}"}:
                reg = re.compile(f)
                cols = [c for c in columns if reg.search(c)]
            elif f in columns:
                # No regular expression.
                cols = [f]
            else:
                continue
            set_cols |= set(cols)
        assert (
            can_be_empty or set_cols
        ), f"Filters {filters} returns an empty set from {columns}"
        return sorted(set_cols)

    def _initialize_columns(self):
        keys = self._filter_column(self._keys, self.data.columns)
        self.values = self._filter_column(self._values, self.data.columns)
        self.ignored = self._filter_column(self._ignored, self.data.columns, True)
        assert (
            self._time in self.data.columns
        ), f"Column {self._time} not found in {pprint.pformat(sorted(self.data.columns))}"
        ignored_keys = set(self.ignored) & set(keys)
        ignored_values = set(self.ignored) & set(self.values)
        self.keys_no_time = [c for c in keys if c not in ignored_keys]
        self.values = [c for c in self.values if c not in ignored_values]
        self.ignored_keys = sorted(ignored_keys)
        self.ignored_values = sorted(ignored_values)
        self.time = self._time
        self.keys_time = [self.time, *[c for c in keys if c not in ignored_keys]]

    def __str__(self) -> str:
        "usual"
        return str(self.data) if hasattr(self, "data") else str(self._data)

    def view(
        self,
        view_def: Union[str, CubeViewDef],
        return_view_def: bool = False,
        verbose: int = 0,
    ) -> Union[pandas.DataFrame, Tuple[pandas.DataFrame, CubeViewDef]]:
        """
        Returns a dataframe, a pivot view.
        `key_index` determines the index, the other key columns determines
        the columns. If `ignore_unique` is True, every columns with a unique value
        is removed.

        :param view_def: view definition
        :param return_view_def: returns the view as well
        :param verbose: verbosity level
        :return: dataframe
        """
        assert isinstance(
            view_def, CubeViewDef
        ), f"view_def should be a CubeViewDef, got {type(view_def)}: {view_def!r} instead"
        if verbose:
            print(f"[CubeLogs.view] -- start view {view_def.name!r}: {view_def}")
        key_agg = (
            self._filter_column(view_def.key_agg, self.keys_time) if view_def.key_agg else []
        )
        set_key_agg = set(key_agg)
        assert set_key_agg <= set(self.keys_time), (
            f"view_def.name={view_def.name!r}, "
            f"non existing keys in key_agg {set_key_agg - set(self.keys_time)}",
            f"keys={sorted(self.keys_time)}",
        )

        values = self._filter_column(view_def.values, self.values)
        assert set(values) <= set(self.values), (
            f"view_def.name={view_def.name!r}, "
            f"non existing columns in values {set(values) - set(self.values)}, "
            f"values={sorted(self.values)}"
        )

        # aggregation
        if key_agg:
            final_stack = True
            key_index = [
                c
                for c in self._filter_column(view_def.key_index, self.keys_time)
                if c not in set_key_agg
            ]
            keys_no_agg = [c for c in self.keys_time if c not in set_key_agg]
            if verbose:
                print(f"[CubeLogs.view] aggregation of {set_key_agg}")
                print(f"[CubeLogs.view] groupby {keys_no_agg}")

            data_red = self.data[[*keys_no_agg, *values]]
            assert set(key_index) <= set(data_red.columns), (
                f"view_def.name={view_def.name!r}, "
                f"nnable to find {set(key_index) - set(data_red.columns)}, "
                f"key_agg={key_agg}, keys_no_agg={keys_no_agg},\n--\n"
                f"selected={pprint.pformat(sorted(data_red.columns))},\n--\n"
                f"keys={pprint.pformat(sorted(self.keys_time))}"
            )
            grouped_data = data_red.groupby(keys_no_agg, as_index=True, dropna=False)
            if callable(view_def.agg_args):
                agg_kwargs = view_def.agg_kwargs or {}
                agg_args = ({c: view_def.agg_args(c) for c in values},)
            else:
                agg_args = view_def.agg_args  # type: ignore[assignment]
                agg_kwargs = view_def.agg_kwargs or {}
            data = grouped_data.agg(*agg_args, **agg_kwargs)
            if view_def.agg_multi:
                append = []
                for k, f in view_def.agg_multi.items():
                    cv = grouped_data.apply(f, include_groups=False)
                    append.append(cv.to_frame(k))
                data = pandas.concat([data, *append], axis=1)
            set_all_keys = set(keys_no_agg)
            values = list(data.columns)
            data = data.reset_index(drop=False)
        else:
            key_index = self._filter_column(view_def.key_index, self.keys_time)
            if verbose:
                print(f"[CubeLogs.view] no aggregation, index={key_index}")
            data = self.data[[*self.keys_time, *values]]
            set_all_keys = set(self.keys_time)
            final_stack = False

        assert set(key_index) <= set_all_keys, (
            f"view_def.name={view_def.name!r}, "
            f"Non existing keys in key_index {set(key_index) - set_all_keys}"
        )

        # remove unnecessary column
        set_key_columns = {
            c for c in self.keys_time if c not in key_index and c not in set(key_agg)
        }
        key_index0 = key_index
        if view_def.ignore_unique:
            unique = {
                k for k, v in self.values_for_key.items() if k in set_all_keys and len(v) <= 1
            }
            keep_anyway = (
                set(view_def.keep_columns_in_index)
                if view_def.keep_columns_in_index
                else set()
            )
            key_index = [k for k in key_index if k not in unique or k in keep_anyway]
            key_columns = [k for k in set_key_columns if k not in unique or k in keep_anyway]
            if verbose:
                print(f"[CubeLogs.view] unique={unique}, keep_anyway={keep_anyway}")
                print(
                    f"[CubeLogs.view] columns with unique values "
                    f"{set(key_index0) - set(key_index)}"
                )
        else:
            if verbose:
                print("[CubeLogs.view] keep all columns")
            key_columns = sorted(set_key_columns)
            unique = set()

        _md = lambda s: {k: v for k, v in self.values_for_key.items() if k in s}  # noqa: E731
        all_cols = set(key_columns) | set(key_index) | set(key_agg) | unique
        assert all_cols == set(self.keys_time), (
            f"view_def.name={view_def.name!r}, "
            f"key_columns + key_index + key_agg + unique != keys, left="
            f"{set(self.keys_time) - all_cols}, "
            f"unique={unique}, index={set(key_index)}, columns={set(key_columns)}, "
            f"agg={set(key_agg)}, keys={set(self.keys_time)}, values={values}"
        )

        # reorder
        if view_def.order:
            subset = self._filter_column(view_def.order, all_cols | {self.time})
            corder = [o for o in view_def.order if o in subset]
            assert set(corder) <= set_key_columns, (
                f"view_def.name={view_def.name!r}, "
                f"non existing columns from order in key_columns "
                f"{set(corder) - set_key_columns}"
            )
            key_columns = [
                *[o for o in corder if o in key_columns],
                *[c for c in key_columns if c not in view_def.order],
            ]
        else:
            corder = None

        if view_def.dropna:
            data, key_index, key_columns, values = self._dropna(  # type: ignore[assignment]
                data,
                key_index,
                key_columns,
                values,
                keep_columns_in_index=view_def.keep_columns_in_index,
            )
        if view_def.ignore_columns:
            if verbose:
                print(f"[CubeLogs.view] ignore_columns {view_def.ignore_columns}")
            data = data.drop(view_def.ignore_columns, axis=1)
            seti = set(view_def.ignore_columns)
            if view_def.keep_columns_in_index:
                seti -= set(view_def.keep_columns_in_index)
            key_index = [c for c in key_index if c not in seti]
            key_columns = [c for c in key_columns if c not in seti]
            values = [c for c in values if c not in seti]

        # final verification
        if verbose:
            print(f"[CubeLogs.view] key_index={key_index}")
            print(f"[CubeLogs.view] key_columns={key_columns}")
        g = data[[*key_index, *key_columns]].copy()
        g["count"] = 1
        r = g.groupby([*key_index, *key_columns], dropna=False).sum()
        not_unique = r[r["count"] > 1]
        assert not_unique.shape[0] == 0, (
            f"view_def.name={view_def.name!r}, "
            f"unable to run the pivot with index={sorted(key_index)}, "
            f"key={sorted(key_columns)}, key_agg={key_agg}, values={sorted(values)}, "
            f"columns={sorted(data.columns)}, ignored={view_def.ignore_columns}, "
            f"not unique={set(data.columns) - unique}"
            f"\n--\n{not_unique.head()}"
        )

        # pivot
        if verbose:
            print(f"[CubeLogs.view] values={values}")
        if key_index:
            piv = data.pivot(index=key_index[::-1], columns=key_columns, values=values)
        else:
            # pivot does return the same rank with it is empty.
            # Let's add arficially one
            data = data.copy()
            data["ALL"] = "ALL"
            piv = data.pivot(index=["ALL"], columns=key_columns, values=values)
        if isinstance(piv, pandas.Series):
            piv = piv.to_frame(name="series")
        names = list(piv.columns.names)
        assert (
            "METRICS" not in names
        ), f"Not implemented when a level METRICS already exists {names!r}"
        names[0] = "METRICS"
        piv.columns = piv.columns.set_names(names)
        if final_stack:
            piv = piv.stack("METRICS", future_stack=True)
        if view_def.transpose:
            piv = piv.T
        if isinstance(piv, pandas.Series):
            piv = piv.to_frame("VALUE")
        piv.sort_index(inplace=True)

        if isinstance(piv.columns, pandas.MultiIndex):
            if corder:
                # reorder the levels for the columns with the view definition
                new_corder = [c for c in corder if c in piv.columns.names]
                new_names = [
                    *[c for c in piv.columns.names if c not in new_corder],
                    *new_corder,
                ]
                piv.columns = piv.columns.reorder_levels(new_names)
            elif self.time in piv.columns.names:
                # put time at the end
                new_names = list(piv.columns.names)
                ind = new_names.index(self.time)
                if ind < len(new_names) - 1:
                    del new_names[ind]
                    new_names.append(self.time)
                    piv.columns = piv.columns.reorder_levels(new_names)

        if view_def.no_index:
            piv = piv.reset_index(drop=False)
        else:
            piv.sort_index(inplace=True, axis=1)

        if verbose:
            print(f"[CubeLogs.view] levels {piv.index.names}, {piv.columns.names}")
            print(f"[CubeLogs.view] -- done view {view_def.name!r}")
        return (piv, view_def) if return_view_def else piv

    def _dropna(
        self,
        data: pandas.DataFrame,
        key_index: Sequence[str],
        key_columns: Sequence[str],
        values: Sequence[str],
        keep_columns_in_index: Optional[Sequence[str]] = None,
    ) -> Tuple[pandas.DataFrame, Sequence[str], Sequence[str], Sequence[str]]:
        set_keep_columns_in_index = (
            set(keep_columns_in_index) if keep_columns_in_index else set()
        )
        v = data[values]
        new_data = data[~v.isnull().all(1)]
        if data.shape == new_data.shape:
            return data, key_index, key_columns, values
        new_data = new_data.copy()
        new_key_index = []
        for c in key_index:
            if c in set_keep_columns_in_index:
                new_key_index.append(c)
                continue
            v = new_data[c]
            sv = set(v.dropna())
            if len(sv) > 1 or (v.isna().max() and len(sv) > 0):
                new_key_index.append(c)
        new_key_columns = []
        for c in key_columns:
            if c in set_keep_columns_in_index:
                new_key_columns.append(c)
                continue
            v = new_data[c]
            sv = set(v.dropna())
            if len(sv) > 1 or (v.isna().max() and len(sv) > 0):
                new_key_columns.append(c)
        for c in set(key_index) | set(key_columns):
            s = new_data[c]
            if s.isna().max():
                if pandas.api.types.is_numeric_dtype(s):
                    min_v = s.dropna().min()
                    assert (
                        min_v >= 0
                    ), f"Unable to replace nan values in column {c!r}, min_v={min_v}"
                    new_data[c] = s.fillna(-1)
                else:
                    new_data[c] = s.fillna("NAN")
        return new_data, new_key_index, new_key_columns, values

    def describe(self) -> pandas.DataFrame:
        """Basic description of all variables."""
        rows = []
        for name in self.data.columns:
            values = self.data[name]
            dtype = values.dtype
            nonan = values.dropna()
            obs = dict(
                name=name,
                dtype=str(dtype),
                missing=len(values) - len(nonan),
                kind=(
                    "time"
                    if name == self.time
                    else (
                        "keys"
                        if name in self.keys_no_time
                        else (
                            "values"
                            if name in self.values
                            else ("ignored" if name in self.ignored else "unused")
                        )
                    )
                ),
            )
            if len(nonan) > 0:
                obs.update(dict(count=len(nonan)))
                if is_numeric_dtype(nonan):
                    obs.update(
                        dict(
                            min=nonan.min(),
                            max=nonan.max(),
                            mean=nonan.mean(),
                            sum=nonan.sum(),
                            n_values=len(set(nonan)),
                        )
                    )
                elif obs["kind"] == "time":
                    unique = set(nonan)
                    obs["n_values"] = len(unique)
                    o = dict(
                        min=str(nonan.min()),
                        max=str(nonan.max()),
                        n_values=len(set(nonan)),
                    )
                    o["values"] = f"{o['min']} - {o['max']}"
                    obs.update(o)
                else:
                    unique = set(nonan)
                    obs["n_values"] = len(unique)
                    if len(unique) < 20:
                        obs["values"] = ",".join(map(str, sorted(unique)))
            rows.append(obs)
        return pandas.DataFrame(rows).set_index("name")

    def to_excel(
        self,
        output: str,
        views: Union[Sequence[str], Dict[str, Union[str, CubeViewDef]]],
        main: Optional[str] = "main",
        raw: Optional[str] = "raw",
        verbose: int = 0,
        csv: Optional[Sequence[str]] = None,
    ):
        """
        Creates an excel file with a list of view.

        :param output: output file to create
        :param views: sequence or dictionary of views to append
        :param main: add a page with statitcs on all variables
        :param raw: add a page with the raw data
        :param csv: views to dump as csv files (same name as outputs + view naw)
        :param verbose: verbosity
        """
        if verbose:
            print(f"[CubeLogs.to_excel] create Excel file {output}, shape={self.shape}")
        views = {k: k for k in views} if not isinstance(views, dict) else views
        f_highlights = {}
        plots = []
        with pandas.ExcelWriter(output, engine="openpyxl") as writer:
            if main:
                assert main not in views, f"{main!r} is duplicated in views {sorted(views)}"
                df = self.describe().sort_values("name")
                if verbose:
                    print(f"[CubeLogs.to_excel] add sheet {main!r} with shape {df.shape}")
                df.to_excel(writer, sheet_name=main, freeze_panes=(1, 1))

            for name, view in views.items():
                if view is None:
                    continue
                df, tview = self.view(view, return_view_def=True, verbose=max(verbose - 1, 0))
                if tview is None:
                    continue
                memory = df.memory_usage(deep=True).sum()
                if verbose:
                    print(
                        f"[CubeLogs.to_excel] add sheet {name!r} with shape "
                        f"{df.shape} ({memory} bytes), index={df.index.names}, "
                        f"columns={df.columns.names}"
                    )
                if self.time in df.columns.names:
                    # Let's convert the time into str
                    fr = df.columns.to_frame()
                    if is_datetime64_any_dtype(fr[self.time]):
                        dt = fr[self.time]
                        has_time = (dt != dt.dt.normalize()).any()
                        sdt = dt.apply(
                            lambda t, has_time=has_time: t.strftime(
                                "%Y-%m-%dT%H-%M-%S" if has_time else "%Y-%m-%d"
                            )
                        )
                        fr[self.time] = sdt
                        df.columns = pandas.MultiIndex.from_frame(fr)
                if csv and name in csv:
                    name_csv = f"{output}.{name}.csv"
                    if verbose:
                        print(f"[CubeLogs.to_excel] saving sheet {name!r} in {name_csv!r}")
                    df.reset_index(drop=False).to_csv(f"{output}.{name}.csv", index=False)

                if memory > 2**22:
                    msg = (
                        f"[CubeLogs.to_excel] skipping {name!r}, "
                        f"too big for excel with {memory} bytes"
                    )
                    if verbose:
                        print(msg)
                    else:
                        warnings.warn(msg, category=RuntimeWarning, stacklevel=0)
                else:
                    df.to_excel(
                        writer,
                        sheet_name=name,
                        freeze_panes=(df.columns.nlevels + df.index.nlevels, df.index.nlevels),
                    )
                    f_highlights[name] = tview.f_highlight
                    if tview.plots:
                        plots.append(
                            CubePlot(
                                df,
                                kind="line",
                                orientation="row",
                                split=True,
                                timeseries=self.time,
                            )
                            if self.time in df.columns.names
                            else CubePlot(df, kind="barh", orientation="row", split=True)
                        )
            if raw:
                assert main not in views, f"{main!r} is duplicated in views {sorted(views)}"
                # Too long.
                # self._apply_excel_style(raw, writer, self.data)
                if csv and "raw" in csv:
                    df.reset_index(drop=False).to_csv(f"{output}.raw.csv", index=False)
                memory = df.memory_usage(deep=True).sum()
                if memory > 2**22:
                    msg = (
                        f"[CubeLogs.to_excel] skipping 'raw', "
                        f"too big for excel with {memory} bytes"
                    )
                    if verbose:
                        print(msg)
                    else:
                        warnings.warn(msg, category=RuntimeWarning, stacklevel=0)
                else:
                    if verbose:
                        print(f"[CubeLogs.to_excel] add sheet 'raw' with shape {self.shape}")
                    self.data.to_excel(
                        writer, sheet_name="raw", freeze_panes=(1, 1), index=True
                    )

            if plots:
                from openpyxl.drawing.image import Image

                if verbose:
                    print(f"[CubeLogs.to_excel] plots {len(plots)} plots")
                sheet = writer.book.create_sheet("plots")
                pos = 0
                empty_row = 1
                times = self.data[self.time].dropna()
                mini, maxi = times.min(), times.max()
                title_suffix = (str(mini) if mini == maxi else f"{mini}-{maxi}").replace(
                    " 00:00:00", ""
                )
                for plot in plots:
                    imgs = plot.to_images(
                        verbose=verbose, merge=True, title_suffix=title_suffix
                    )
                    for img in imgs:
                        y = (pos // 2) * 16
                        loc = f"A{y}" if pos % 2 == 0 else f"M{y}"
                        sheet.add_image(Image(io.BytesIO(img)), loc)
                        if verbose:
                            no = f"{output}.png"
                            print(f"[CubeLogs.to_excel] dump graphs into {no!r}")
                            with open(no, "wb") as f:
                                f.write(img)
                        pos += 1
                    empty_row += len(plots) + 2

            if verbose:
                print(f"[CubeLogs.to_excel] applies style to {output!r}")
            apply_excel_style(writer, f_highlights)  # type: ignore[arg-type]
            if verbose:
                print(f"[CubeLogs.to_excel] done with {len(views)} views")


class CubeLogsPerformance(CubeLogs):
    """
    Processes logs coming from experiments.
    """

    def __init__(
        self,
        data: Any,
        time: str = "DATE",
        keys: Sequence[str] = (
            "^version_.*",
            "^model_.*",
            "device",
            "opt_patterns",
            "suite",
            "memory_peak",
            "machine",
            "exporter",
            "dynamic",
            "rtopt",
            "dtype",
            "device",
            "architecture",
        ),
        values: Sequence[str] = (
            "^time_.*",
            "^disc.*",
            "^ERR_.*",
            "CMD",
            "^ITER",
            "^onnx_.*",
            "^op_onnx_.*",
            "^peak_gpu_.*",
        ),
        ignored: Sequence[str] = ("version_python",),
        recent: bool = True,
        formulas: Optional[
            Union[
                Sequence[str],
                Dict[str, Union[str, Callable[[pandas.DataFrame], pandas.Series]]],
            ]
        ] = (
            "speedup",
            "bucket[speedup]",
            "ERR1",
            "n_models",
            "n_model_eager",
            "n_model_running",
            "n_model_acc01",
            "n_model_acc001",
            "n_model_dynamic",
            "n_model_pass",
            "n_model_faster",
            "n_model_faster2x",
            "n_model_faster3x",
            "n_model_faster4x",
            "n_node_attention",
            "n_node_control_flow",
            "n_node_scatter",
            "n_node_function",
            "n_node_initializer",
            "n_node_initializer_small",
            "n_node_constant",
            "n_node_shape",
            "n_node_expand",
            "onnx_n_nodes_no_cst",
            "peak_gpu_torch",
            "peak_gpu_nvidia",
            "time_export_unbiased",
        ),
        fill_missing: Optional[Sequence[Tuple[str, Any]]] = (("model_attn_impl", "eager"),),
        keep_last_date: bool = False,
    ):
        super().__init__(
            data=data,
            time=time,
            keys=keys,
            values=values,
            ignored=ignored,
            recent=recent,
            formulas=formulas,
            fill_missing=fill_missing,
            keep_last_date=keep_last_date,
        )

    def _process_formula(
        self, formula: Union[str, Callable[[pandas.DataFrame], pandas.Series]]
    ) -> Callable[[pandas.DataFrame], pandas.Series]:
        """
        Processes a formula, converting it into a function.

        :param formula: a formula string
        :return: a function
        """
        if callable(formula):
            return formula
        assert isinstance(
            formula, str
        ), f"Unexpected type for formula {type(formula)}: {formula!r}"

        def gdf(df, cname, default_value=np.nan):
            if cname in df.columns:
                return df[cname]
            return pandas.Series(default_value, index=df.index)

        def ghas_value(df, cname):
            if cname not in df.columns:
                return pandas.Series(np.nan, index=df.index)
            isna = df[cname].isna()
            return pandas.Series(np.where(isna, np.nan, 1.0), index=df.index)

        def gpreserve(df, cname, series):
            if cname not in df.columns:
                return pandas.Series(np.nan, index=df.index)
            isna = df[cname].isna()
            return pandas.Series(np.where(isna, np.nan, series), index=df.index).astype(float)

        if formula == "speedup":
            columns = set(self._filter_column(["^time_.*"], self.data.columns))
            assert "time_latency" in columns and "time_latency_eager" in columns, (
                f"Unable to apply formula {formula!r}, with columns\n"
                f"{pprint.pformat(sorted(columns))}"
            )
            return lambda df: df["time_latency_eager"] / df["time_latency"]

        if formula == "bucket[speedup]":
            columns = set(self._filter_column(["^time_.*", "speedup"], self.data.columns))
            assert "speedup" in columns, (
                f"Unable to apply formula {formula!r}, with columns\n"
                f"{pprint.pformat(sorted(columns))}"
            )
            # return lambda df: df["time_latency_eager"] / df["time_latency"]
            return lambda df: pandas.cut(
                df["speedup"], bins=BUCKET_SCALES, right=False, duplicates="raise"
            )

        if formula == "ERR1":
            columns = set(self._filter_column(["^ERR_.*"], self.data.columns))
            if not columns:
                return lambda df: np.nan

            def first_err(df: pandas.DataFrame) -> pandas.Series:
                ordered = [
                    c
                    for c in [
                        "ERR_timeout",
                        "ERR_load",
                        "ERR_feeds",
                        "ERR_warmup_eager",
                        "ERR_export",
                        "ERR_ort",
                        "ERR_warmup",
                        # "ERR_std",
                        # "ERR_crash",
                        # "ERR_stdout",
                    ]
                    if c in df.columns
                ]
                res = None
                for c in ordered:
                    if res is None:
                        res = df[c].fillna("")
                    else:
                        res = pandas.Series(np.where(res != "", res, df[c].fillna("")))
                return res

            return first_err

        if formula.startswith("n_"):
            lambdas = dict(
                n_models=lambda df: ghas_value(df, "model_name"),
                n_model_eager=lambda df: ghas_value(df, "time_latency_eager"),
                n_model_running=lambda df: ghas_value(df, "time_latency"),
                n_model_acc01=lambda df: gpreserve(
                    df, "discrepancies_abs", (gdf(df, "discrepancies_abs") <= 0.1)
                ),
                n_model_acc001=lambda df: gpreserve(
                    df, "discrepancies_abs", gdf(df, "discrepancies_abs") <= 0.01
                ),
                n_model_dynamic=lambda df: gpreserve(
                    df,
                    "discrepancies_dynamic_abs",
                    (gdf(df, "discrepancies_dynamic_abs") <= 0.1),
                ),
                n_model_pass=lambda df: gpreserve(
                    df,
                    "time_latency",
                    (gdf(df, "discrepancies_abs", np.inf) < 0.1)
                    & (gdf(df, "time_latency_eager") > gdf(df, "time_latency", np.inf) * 0.98),
                ),
                n_model_faster=lambda df: gpreserve(
                    df,
                    "time_latency",
                    gdf(df, "time_latency_eager") > gdf(df, "time_latency", np.inf) * 0.98,
                ),
                n_model_faster2x=lambda df: gpreserve(
                    df,
                    "time_latency",
                    gdf(df, "time_latency_eager") > gdf(df, "time_latency", np.inf) * 1.98,
                ),
                n_model_faster3x=lambda df: gpreserve(
                    df,
                    "time_latency",
                    gdf(df, "time_latency_eager") > gdf(df, "time_latency", np.inf) * 2.98,
                ),
                n_model_faster4x=lambda df: gpreserve(
                    df,
                    "time_latency",
                    gdf(df, "time_latency_eager") > gdf(df, "time_latency", np.inf) * 3.98,
                ),
                n_node_attention=lambda df: gpreserve(
                    df,
                    "op_onnx_com.microsoft_Attention",
                    gdf(df, "op_onnx_com.microsoft_Attention")
                    + gdf(df, "op_onnx_com.microsoft_MultiHeadAttention"),
                ),
                n_node_control_flow=lambda df: gpreserve(
                    df,
                    "op_onnx__If",
                    (
                        gdf(df, "op_onnx__If", 0)
                        + gdf(df, "op_onnx__Scan", 0)
                        + gdf(df, "op_onnx__Loop", 0)
                    ),
                ),
                n_node_scatter=lambda df: gpreserve(
                    df,
                    "op_onnx__ScatterND",
                    gdf(df, "op_onnx__ScatterND", 0) + gdf(df, "op_onnx__ScatterElements", 0),
                ),
                n_node_function=lambda df: gpreserve(
                    df, "onnx_n_functions", gdf(df, "onnx_n_functions")
                ),
                n_node_initializer_small=lambda df: gpreserve(
                    df, "op_onnx_initializer_small", gdf(df, "op_onnx_initializer_small")
                ),
                n_node_initializer=lambda df: gpreserve(
                    df, "onnx_n_initializer", gdf(df, "onnx_n_initializer")
                ),
                n_node_constant=lambda df: gpreserve(
                    df, "op_onnx__Constant", gdf(df, "op_onnx__Constant")
                ),
                n_node_shape=lambda df: gpreserve(
                    df, "op_onnx__Shape", gdf(df, "op_onnx__Shape")
                ),
                n_node_expand=lambda df: gpreserve(
                    df, "op_onnx__Expand", gdf(df, "op_onnx__Expand")
                ),
            )
            assert (
                formula in lambdas
            ), f"Unexpected formula={formula!r}, should be in {sorted(lambdas)}"
            return lambdas[formula]

        if formula == "onnx_n_nodes_no_cst":
            return lambda df: gdf(df, "onnx_n_nodes", 0) - gdf(
                df, "op_onnx__Constant", 0
            ).fillna(0)
        if formula == "peak_gpu_torch":
            return lambda df: gdf(df, "mema_gpu_5_after_export") - gdf(df, "mema_gpu_4_reset")
        if formula == "peak_gpu_nvidia":
            return (
                lambda df: (gdf(df, "memory_gpu0_peak") - gdf(df, "memory_gpu0_begin")) * 2**20
            )
        if formula == "time_export_unbiased":

            def unbiased_export(df):
                if "time_warmup_first_iteration" not in df.columns:
                    return pandas.Series(np.nan, index=df.index)
                return pandas.Series(
                    np.where(
                        df["exporter"] == "inductor",
                        df["time_warmup_first_iteration"] + df["time_export_success"],
                        df["time_export_success"],
                    ),
                    index=df.index,
                )

            return lambda df: gpreserve(df, "time_warmup_first_iteration", unbiased_export(df))

        raise ValueError(
            f"Unexpected formula {formula!r}, available columns are\n"
            f"{pprint.pformat(sorted(self.data.columns))}"
        )

    def view(
        self,
        view_def: Optional[Union[str, CubeViewDef]],
        return_view_def: bool = False,
        verbose: int = 0,
    ) -> Union[
        Optional[pandas.DataFrame], Tuple[Optional[pandas.DataFrame], Optional[CubeViewDef]]
    ]:
        """
        Returns a dataframe, a pivot view.

        If view_def is a string, it is replaced by a prefined view.

        :param view_def: view definition or a string
        :param return_view_def: returns the view definition as well
        :param verbose: verbosity level
        :return: dataframe or a couple (dataframe, view definition),
            both of them can be one if view_def cannot be interpreted
        """
        assert view_def is not None, "view_def is None, this is not allowed."
        if isinstance(view_def, str):
            view_def = self.make_view_def(view_def)
            if view_def is None:
                return (None, None) if return_view_def else None
        return super().view(view_def, return_view_def=return_view_def, verbose=verbose)

    def make_view_def(self, name: str) -> Optional[CubeViewDef]:
        """
        Returns a view definition.

        :param name: name of the view
        :return: a CubeViewDef or None if name does not make sense

        Available views:

        * **agg-suite:** aggregation per suite
        * **disc:** discrepancies
        * **speedup:** speedup
        * **bucket_speedup:** speedup in buckets
        * **time:** latency
        * **time_export:** time to export
        * **counts:** status, running, faster, has control flow, ...
        * **err:** important errors
        * **cmd:** command lines
        * **raw-short:** raw data without all the unused columns
        """
        fs = ["suite", "model_suite", "task", "model_name", "model_task"]
        index_cols = self._filter_column(fs, self.keys_time)
        assert index_cols, (
            f"No index columns found for {fs!r} in "
            f"{pprint.pformat(sorted(self.keys_time))}"
        )
        index_cols = [c for c in fs if c in set(index_cols)]

        f_speedup = lambda x: (  # noqa: E731
            CubeViewDef.HighLightKind.NONE
            if not isinstance(x, (float, int))
            else (
                CubeViewDef.HighLightKind.RED
                if x < 0.9
                else (
                    CubeViewDef.HighLightKind.GREEN
                    if x > 1.1
                    else CubeViewDef.HighLightKind.NONE
                )
            )
        )
        f_disc = lambda x: (  # noqa: E731
            CubeViewDef.HighLightKind.NONE
            if not isinstance(x, (float, int))
            else (
                CubeViewDef.HighLightKind.RED
                if x > 0.1
                else (
                    CubeViewDef.HighLightKind.GREEN
                    if x < 0.01
                    else CubeViewDef.HighLightKind.NONE
                )
            )
        )
        f_bucket = lambda x: (  # noqa: E731
            CubeViewDef.HighLightKind.NONE
            if not isinstance(x, str)
            else (
                CubeViewDef.HighLightKind.RED
                if x in {"[-inf, 0.8)", "[0.8, 0.9)", "[0.9, 0.95)"}
                else (
                    CubeViewDef.HighLightKind.NONE
                    if x in {"[0.95, 0.98)", "[0.98, 1.02)", "[1.02, 1.05)"}
                    else (
                        CubeViewDef.HighLightKind.GREEN
                        if "[" in x
                        else CubeViewDef.HighLightKind.NONE
                    )
                )
            )
        )

        def mean_weight(gr):
            weight = gr["time_latency_eager"]
            x = gr["speedup"]
            if x.shape[0] == 0:
                return np.nan
            div = weight.sum()
            if div > 0:
                return (x * weight).sum() / div
            return np.nan

        def mean_geo(gr):
            x = gr["speedup"]
            return np.exp(np.log(x.dropna()).mean())

        order = ["model_attn_impl", "exporter", "opt_patterns", "DATE"]
        implemented_views = {
            "agg-suite": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(
                    [
                        "TIME_ITER",
                        "speedup",
                        "time_latency",
                        "time_latency_eager",
                        "time_export_success",
                        "time_export_unbiased",
                        "^n_.*",
                        "target_opset",
                        "onnx_filesize",
                        "onnx_weight_size_torch",
                        "onnx_weight_size_proto",
                        "onnx_n_nodes",
                        "onnx_n_nodes_no_cst",
                        "op_onnx__Constant",
                        "peak_gpu_torch",
                        "peak_gpu_nvidia",
                    ],
                    self.values,
                ),
                ignore_unique=True,
                key_agg=["model_name", "task", "model_task"],
                agg_args=lambda column_name: "sum" if column_name.startswith("n_") else "mean",
                agg_multi={"speedup_weighted": mean_weight, "speedup_geo": mean_geo},
                keep_columns_in_index=["suite"],
                name="agg-suite",
                order=order,
            ),
            "agg-all": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(
                    [
                        "TIME_ITER",
                        "speedup",
                        "time_latency",
                        "time_latency_eager",
                        "time_export_success",
                        "time_export_unbiased",
                        "^n_.*",
                        "target_opset",
                        "onnx_filesize",
                        "onnx_weight_size_torch",
                        "onnx_weight_size_proto",
                        "onnx_n_nodes",
                        "onnx_n_nodes_no_cst",
                        "peak_gpu_torch",
                        "peak_gpu_nvidia",
                    ],
                    self.values,
                ),
                ignore_unique=True,
                key_agg=["model_name", "task", "model_task", "suite"],
                agg_args=lambda column_name: "sum" if column_name.startswith("n_") else "mean",
                agg_multi={"speedup_weighted": mean_weight, "speedup_geo": mean_geo},
                name="agg-all",
                order=order,
                plots=True,
            ),
            "disc": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["discrepancies_abs"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                f_highlight=f_disc,
                name="disc",
                order=order,
            ),
            "speedup": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["speedup"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                f_highlight=f_speedup,
                name="speedup",
                order=order,
            ),
            "counts": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["^n_.*"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="counts",
                order=order,
            ),
            "peak-gpu": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["^peak_gpu_.*"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="peak-gpu",
                order=order,
            ),
            "time": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(
                    ["time_latency", "time_latency_eager"], self.values
                ),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="time",
                order=order,
            ),
            "time_export": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["time_export_unbiased"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="time_export",
                order=order,
            ),
            "err": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(
                    ["ERR1", "ERR_timeout", "ERR_export", "ERR_crash"], self.values
                ),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="err",
                order=order,
            ),
            "bucket-speedup": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(["bucket[speedup]"], self.values),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="bucket-speedup",
                f_highlight=f_bucket,
                order=order,
            ),
            "onnx": lambda: CubeViewDef(
                key_index=index_cols,
                values=self._filter_column(
                    [
                        "onnx_filesize",
                        "onnx_n_nodes",
                        "onnx_n_nodes_no_cst",
                        "onnx_weight_size_proto",
                        "onnx_weight_size_torch",
                        "op_onnx_initializer_small",
                    ],
                    self.values,
                ),
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="onnx",
                order=order,
            ),
            "raw-short": lambda: CubeViewDef(
                key_index=self.keys_time,
                values=[c for c in self.values if c not in {"ERR_std", "ERR_stdout"}],
                ignore_unique=False,
                keep_columns_in_index=["suite"],
                name="raw-short",
                no_index=True,
            ),
        }

        cmd_col = self._filter_column(["CMD"], self.values, can_be_empty=True)
        if cmd_col:
            implemented_views["cmd"] = lambda: CubeViewDef(
                key_index=index_cols,
                values=cmd_col,
                ignore_unique=True,
                keep_columns_in_index=["suite"],
                name="cmd",
                order=order,
            )

        assert name in implemented_views or name in {"cmd"}, (
            f"Unknown view {name!r}, expected a name in {sorted(implemented_views)},"
            f"\n--\nkeys={pprint.pformat(sorted(self.keys_time))}, "
            f"\n--\nvalues={pprint.pformat(sorted(self.values))}"
        )
        if name not in implemented_views:
            return None
        return implemented_views[name]()

    def post_load_process_piece(
        self, df: pandas.DataFrame, unique: bool = False
    ) -> pandas.DataFrame:
        df = super().post_load_process_piece(df, unique=unique)
        if unique:
            return df
        cols = self._filter_column(self._keys, df)
        res = None
        for c in cols:
            if df[c].isna().any():
                # Missing values for keys are not supposed to happen.
                uniq = set(df[c].dropna())
                if len(uniq) == 1:
                    if res is None:
                        res = df.copy()
                    res[c] = res[c].fillna(uniq.pop())
        return df if res is None else res
